"""
Watchlist load, backed by a dated file in watchlist/, uploaded manually
each day (GitHub web UI or app) before market open. Accepts EITHER format,
whichever you happen to upload that day:

    watchlist/<YYYY-MM-DD>.json
    watchlist/<YYYY-MM-DD>.xlsx

using the US Eastern calendar date (so it lines up with the trading session
regardless of what timezone/date it is where you upload from). If both
exist for the same day, the .json is used (checked first). See
watchlist/README.md for the exact format of each.

Both scanner_us.py and institutional_scanner_us.py re-read this at the top
of every polling loop iteration, so if you fix a typo mid-session and
re-upload the file, it takes effect on the next poll -- no restart needed.
"""

import json
import logging
import os
from datetime import datetime
from zoneinfo import ZoneInfo

WATCHLIST_DIR = "watchlist"
US_TZ = ZoneInfo("America/New_York")

# Excel column headers this looks for (case-insensitive). Only "symbol" is
# required -- every other column is optional and can be left blank.
XLSX_COLUMNS = {
    "symbol": ("symbol", "ticker"),
    "resistance": ("resistance",),
    "support": ("support",),
    "pivot_level": ("pivot_level", "pivot level", "pivot"),
    "setup": ("setup",),
    "description": ("description", "note", "notes"),
}

_warned_missing_today = False  # only log the "no file yet" warning once per process


def _today_base_path() -> str:
    today = datetime.now(US_TZ).strftime("%Y-%m-%d")
    return os.path.join(WATCHLIST_DIR, today)


def _to_float(v):
    if v in (None, "", "null", "None", "-"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_str(v):
    return str(v).strip() if v not in (None, "") else ""


def _load_json(path: str) -> dict:
    with open(path, "r") as f:
        return json.load(f)


def _load_xlsx(path: str) -> dict:
    """One row per symbol. Header row required; column order doesn't
    matter, only the names (case-insensitive, see XLSX_COLUMNS above)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [_to_str(c.value).lower() for c in header_cells]

    def col_index(names):
        for name in names:
            if name in headers:
                return headers.index(name)
        return None

    idx = {key: col_index(names) for key, names in XLSX_COLUMNS.items()}
    if idx["symbol"] is None:
        raise ValueError("No 'Symbol' (or 'Ticker') column found in header row")

    def cell(row, key):
        i = idx[key]
        return row[i] if i is not None and i < len(row) else None

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        symbol = _to_str(cell(row, "symbol")).upper()
        if not symbol:
            continue  # skip blank rows
        result[symbol] = {
            "resistance": _to_float(cell(row, "resistance")),
            "support": _to_float(cell(row, "support")),
            "pivot_level": _to_float(cell(row, "pivot_level")),
            "setup": _to_str(cell(row, "setup")),
            "description": _to_str(cell(row, "description")),
        }
    return result


def load_watchlist() -> dict:
    """Returns {} if today's file doesn't exist yet in either format --
    callers should treat an empty watchlist as "nothing to scan yet" and
    just wait for the next poll (e.g. file uploaded a few minutes late)."""
    global _warned_missing_today
    base = _today_base_path()
    json_path, xlsx_path = f"{base}.json", f"{base}.xlsx"

    try:
        if os.path.exists(json_path):
            data = _load_json(json_path)
        elif os.path.exists(xlsx_path):
            data = _load_xlsx(xlsx_path)
        else:
            if not _warned_missing_today:
                logging.warning(f"No watchlist file found at {json_path} or {xlsx_path} -- "
                                 f"upload today's list (either format) to watchlist/ before market open.")
                _warned_missing_today = True
            return {}
    except (json.JSONDecodeError, OSError, ValueError) as e:
        logging.error(f"Couldn't parse today's watchlist file: {e}")
        return {}

    _warned_missing_today = False
    return data


def parse_watchlist_text(text: str):
    """
    Kept for convenience if you ever want to hand-convert a pasted list into
    JSON locally. Parses one ticker per line, comma or tab separated:
        SYRE,108,98
        CRWD,219,202,
        DFTX,46.17,44.45,49.47
        NVDA
    A ticker alone (no levels) is still watched -- it just won't trigger the
    breakout/undercut-rally/30min-pivot/pivot-cross setups, which need a
    level. The RVOL-based and institutional-flow alerts have no level
    requirement and still fire for it.

    Returns (watchlist_dict, errors).
    """
    result = {}
    errors = []

    for raw_line in text.strip().splitlines():
        line = raw_line.strip()
        if not line or line.startswith("/"):
            continue

        parts = [p.strip() for p in line.replace("\t", ",").split(",")]
        ticker = parts[0].upper()
        if not ticker.isalnum():
            errors.append(f"Skipped invalid ticker: '{raw_line}'")
            continue

        try:
            resistance = _to_float(parts[1]) if len(parts) > 1 else None
            support = _to_float(parts[2]) if len(parts) > 2 else None
            pivot = _to_float(parts[3]) if len(parts) > 3 else None
        except ValueError:
            errors.append(f"Skipped '{raw_line}' -- couldn't parse a number")
            continue

        result[ticker] = {"resistance": resistance, "support": support, "pivot_level": pivot}

    return result, errors
